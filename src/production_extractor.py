import argparse
import csv
import json
import re
import sys
from pathlib import Path

for stream in (sys.stdout, sys.stderr):
    if hasattr(stream, "reconfigure"):
        stream.reconfigure(encoding="utf-8")

SCENE_HEADER_RE = re.compile(r"^## SCENE (\d{3}) \[zone: ([A-Za-z0-9_]+)\]$")
TAG_RE = re.compile(r"^(NARRATION|SPEAKER|SE): (.*)$")
CUT_RE = re.compile(r"^CUT (\d+): (.*)$")

_HALFWIDTH_DIGITS_SYMBOLS = "0123456789%-:/+.,()!?"
_FULLWIDTH_DIGITS_SYMBOLS = "０１２３４５６７８９％－：／＋．，（）！？"
_FULLWIDTH_TABLE = str.maketrans(_HALFWIDTH_DIGITS_SYMBOLS, _FULLWIDTH_DIGITS_SYMBOLS)

# DRAFT値・暫定リスト。固定ルールではない。誤検知・見逃し双方あり得る前提で
# 人間判断を補助するためのフラグ立てに留める（自動変換・自動ブロックはしない）
POLICY_RISK_KEYWORDS = [
    "入浴", "お風呂", "着替え", "排泄", "排せつ", "おむつ", "点滴", "注射", "下着", "裸",
]


def parse_script(text: str):
    lines = text.splitlines()
    if not lines or lines[0].strip() != "---":
        raise ValueError("先頭にメタブロック(---)がありません")
    meta, body_start = _parse_meta_block(lines)
    scenes = _parse_scenes(lines[body_start:])
    return meta, scenes


def _parse_meta_block(lines):
    end = None
    for i in range(1, len(lines)):
        if lines[i].strip() == "---":
            end = i
            break
    if end is None:
        raise ValueError("メタブロックの終端(---)が見つかりません")

    meta = {}
    current_nested_key = None
    for line in lines[1:end]:
        if not line.strip():
            continue
        if line.startswith("  ") and current_nested_key:
            key, value = _split_kv(line.strip())
            meta[current_nested_key][key] = _coerce(value)
            continue
        key, value = _split_kv(line.strip())
        if value == "":
            meta[key] = {}
            current_nested_key = key
        else:
            meta[key] = _coerce(value)
            current_nested_key = None
    return meta, end + 1


def _split_kv(line):
    key, _, value = line.partition(":")
    return key.strip(), value.strip()


def _coerce(value):
    return int(value) if value.isdigit() else value


def _parse_scenes(lines):
    scenes = []
    current = None
    for line in lines:
        header = SCENE_HEADER_RE.match(line)
        if header:
            if current:
                scenes.append(current)
            current = {
                "scene_id": header.group(1),
                "zone_id": header.group(2),
                "narration": "",
                "speaker": "ナレーター",
                "cuts": [],
                "se": "",
            }
            continue
        if current is None:
            continue
        tag = TAG_RE.match(line)
        if tag:
            name, value = tag.group(1), tag.group(2)
            if name == "NARRATION":
                current["narration"] = value
            elif name == "SPEAKER":
                current["speaker"] = value
            elif name == "SE":
                current["se"] = value
            continue
        cut = CUT_RE.match(line)
        if cut:
            current["cuts"].append(cut.group(2))
            continue
        if line.startswith("IMAGE_PROMPT: "):
            current["cuts"].append(line[len("IMAGE_PROMPT: "):])
    if current:
        scenes.append(current)
    return scenes


def load_zone_budgets(genre_path: Path, duration_sec: int) -> dict:
    genre = json.loads(genre_path.read_text(encoding="utf-8"))
    chars_per_min = genre["narration_speed"]["chars_per_min"]
    ratio = genre["narration_speed"]["effective_char_ratio"]
    budgets = {}
    for zone in genre["zones"]:
        zone_duration = max(duration_sec * zone["duration_ratio"], zone.get("min_sec", 0))
        max_sec = zone.get("max_sec")
        if max_sec is not None:
            zone_duration = min(zone_duration, max_sec)
        budgets[zone["zone_id"]] = zone_duration / 60 * chars_per_min * ratio
    return budgets


def validate_total_scenes(meta, scenes, warn):
    expected = meta.get("total_scenes")
    actual = len(scenes)
    if expected is not None and expected != actual:
        warn(f"total_scenes不一致: メタブロック={expected} 実際={actual}")


def validate_zone_char_budgets(meta, scenes, genre_path, warn):
    duration_sec = meta.get("duration_sec")
    if duration_sec is None:
        warn("メタブロックにduration_secが無いためゾーン文字数チェックをスキップ")
        return
    if not genre_path.exists():
        warn(f"genres/risk_sim.jsonが見つからないためゾーン文字数チェックをスキップ: {genre_path}")
        return

    budgets = load_zone_budgets(genre_path, duration_sec)
    actual_chars = {}
    for scene in scenes:
        actual_chars[scene["zone_id"]] = actual_chars.get(scene["zone_id"], 0) + len(scene["narration"])

    for zone_id, budget in budgets.items():
        if budget == 0:
            continue
        actual = actual_chars.get(zone_id, 0)
        diff_ratio = (actual - budget) / budget
        if abs(diff_ratio) > 0.15:
            warn(
                f"{zone_id}: NARRATION文字数が予算から{diff_ratio:+.0%}乖離"
                f"（実測{actual}字 / 予算{budget:.0f}字）"
            )


def validate_cut_pacing(meta, scenes, genre_path: Path, warn):
    duration_sec = meta.get("duration_sec")
    if duration_sec is None:
        return
    if not genre_path.exists():
        return
    genre = json.loads(genre_path.read_text(encoding="utf-8"))
    visual_pacing = genre.get("visual_pacing")
    if not visual_pacing:
        return

    chars_per_min = genre["narration_speed"]["chars_per_min"]
    ratio = genre["narration_speed"]["effective_char_ratio"]
    default_min = visual_pacing.get("min_cut_sec")
    default_max = visual_pacing.get("max_cut_sec")
    zone_overrides = visual_pacing.get("zone_overrides", {})

    for scene in scenes:
        override = zone_overrides.get(scene["zone_id"], {})
        zone_min = override.get("min_cut_sec", default_min)
        zone_max = override.get("max_cut_sec", default_max)

        estimated_sec = len(scene["narration"]) / (chars_per_min * ratio) * 60
        cut_count = len(scene["cuts"]) or 1
        sec_per_cut = estimated_sec / cut_count

        if zone_max is not None and sec_per_cut > zone_max * 1.5:
            warn(
                f"SCENE {scene['scene_id']}: 1カットあたり{sec_per_cut:.1f}秒でカットが粗い"
                f"（上限目安{zone_max}秒の1.5倍を超過）"
            )
        if zone_min is not None and sec_per_cut < zone_min * 0.5 and estimated_sec > zone_min:
            warn(
                f"SCENE {scene['scene_id']}: 1カットあたり{sec_per_cut:.1f}秒でカットが細かすぎる"
                f"（下限目安{zone_min}秒の0.5倍未満）"
            )


def validate_policy_risk(scenes, warn):
    for scene in scenes:
        if any(kw in scene["narration"] for kw in POLICY_RISK_KEYWORDS):
            warn(
                f"[POLICY_CHECK] SCENE {scene['scene_id']}: ポリシーリスクの可能性。"
                f"人間または高位モデルでの確認を推奨"
            )


def validate_image_prompts(scenes, warn):
    for scene in scenes:
        if not scene["cuts"]:
            warn(f"SCENE {scene['scene_id']}: CUTが1つもありません")
            continue
        for i, cut_prompt in enumerate(scene["cuts"], start=1):
            if not cut_prompt.strip():
                warn(f"SCENE {scene['scene_id']} CUT {i}: 画像プロンプトが空です")


def validate_metrics_columns(metrics_path: Path, warn):
    if not metrics_path.exists():
        warn(f"metrics.csvが見つかりません: {metrics_path}")
        return
    with metrics_path.open(encoding="utf-8-sig", newline="") as f:
        header = next(csv.reader(f), [])
    for col in ("ending_pattern", "explanation_layout"):
        if col not in header:
            warn(f"metrics.csvに列 '{col}' が見つかりません")


def _to_fullwidth(text: str) -> str:
    return text.translate(_FULLWIDTH_TABLE)


def write_voice_plain_text(scenes, out_dir: Path):
    text = "\n\n".join(_to_fullwidth(scene["narration"]) for scene in scenes)
    (out_dir / "voice_plain_text.txt").write_text(text, encoding="utf-8")


def write_voice_assignment_csv(scenes, out_dir: Path):
    with (out_dir / "voice_assignment.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["scene_id", "zone_id", "speaker"])
        for scene in scenes:
            writer.writerow([scene["scene_id"], scene["zone_id"], scene["speaker"]])


WORD_RE = re.compile(r"[A-Za-z]+")

# image_prompt本文が抽象図解（キャラクター/ロケーションを意図的に適用しないカット）
# であることを示すマーカー句。DRAFT値・暫定リスト
ABSTRACT_DIAGRAM_MARKERS = [
    "informational illustration",
    "conceptual illustration",
    "editorial illustration",
    "checklist infographic",
]

# 特徴語抽出から除外する一般的な機能語（前置詞・冠詞・接続詞等）。
# これらはlocked_prompt_fragment間でたまたま非対称に出現するだけで、
# キャラクター/ロケーションを識別する特徴語ではないため除外する
STOPWORDS = {
    "the", "and", "for", "are", "but", "not", "all", "any", "can", "had", "has",
    "was", "one", "our", "out", "she", "too", "use", "with", "this", "that",
    "from", "they", "will", "there", "their", "what", "about", "which", "when",
    "make", "just", "into", "over", "after", "also", "been", "being", "both",
    "down", "each", "few", "have", "here", "how", "its", "itself", "more",
    "most", "other", "same", "such", "than", "very", "him", "himself", "who",
    "whom", "these", "those", "were", "off", "own", "only", "under", "until",
    "while", "above", "below", "between", "during", "before", "again",
    "further", "once", "nor", "onto", "upon", "near", "toward", "towards",
    "within", "without", "against", "among", "along", "around", "because",
    "behind", "beneath", "beside", "besides", "despite", "except", "inside",
    "instead", "outside", "per", "plus", "regarding", "since", "though",
    "throughout", "unless", "unlike", "versus", "via", "whereas", "yet",
    # 全カット共通の様式ボイラープレート／照明技法語との衝突を避けるため除外。
    # 例："clean"は"crisp clean linework"（全カット共通の描画スタイル指定）で
    # 毎回出現し、"low"は"low-key ... lighting"（照明トーン指定）で頻出するため、
    # ロケーション判定の特徴語としては機能しない
    "low", "clean",
}


def _extract_words(text: str) -> set:
    return {
        w.lower()
        for w in WORD_RE.findall(text)
        if len(w) >= 3 and w.lower() not in STOPWORDS
    }


def _distinctive_terms(entries: list, id_key: str, fragment_key: str) -> dict:
    """各entryのlocked_prompt_fragmentから、他entryと重複しない特徴語集合を抽出する"""
    word_sets = {e[id_key]: _extract_words(e.get(fragment_key, "")) for e in entries}
    distinctive = {}
    for eid, words in word_sets.items():
        others = set()
        for other_id, other_words in word_sets.items():
            if other_id != eid:
                others |= other_words
        distinctive[eid] = words - others
    return distinctive


def _is_abstract_diagram(text: str) -> bool:
    lowered = text.lower()
    return any(marker in lowered for marker in ABSTRACT_DIAGRAM_MARKERS)


def _classify_cut_references(text: str, char_distinctive: dict, loc_distinctive: dict):
    if _is_abstract_diagram(text):
        return [], None, "抽象図解のためキャラクター/ロケーションのシート参照は不要"

    text_words = _extract_words(text)

    # character_refs: 特徴語が複数（2語以上）含まれていれば一致とみなす
    character_refs = [
        cid for cid, terms in char_distinctive.items() if len(terms & text_words) >= 2
    ]

    # location_ref: 最も一致数の多いロケーションを採用（1語でも一致していれば候補）
    location_ref = None
    best_hits = 0
    for lid, terms in loc_distinctive.items():
        hits = len(terms & text_words)
        if hits > best_hits:
            location_ref, best_hits = lid, hits

    notes_parts = []
    if not character_refs and not location_ref:
        notes_parts.append("キャラクター/ロケーションいずれの特徴語も検出できず（判定不可）")
    elif not location_ref:
        notes_parts.append("地の文に場所描写なし（ロケーション判定不可）")

    lowered = text.lower()
    for label, terms in ROLE_SEARCH_TERMS.items():
        if any(t.lower() in lowered for t in terms):
            notes_parts.append(f"visual_cast未登録の脇役「{label}」が登場（シート参照対象外）")

    return character_refs, location_ref, "；".join(notes_parts)


def _build_sheet_index(visual_sheets_dir: Path) -> dict:
    """visual_sheets配下のPNGファイル名から character_id/location_id -> ファイル名 を作る"""
    index = {}
    if not visual_sheets_dir.exists():
        return index
    for f in sorted(visual_sheets_dir.glob("*.png")):
        m = re.match(r"^([A-Za-z0-9]+)_", f.name)
        if m:
            index[m.group(1)] = f.name
    return index


def write_image_prompts(scenes, blueprint: dict, visual_sheets_dir: Path, out_dir: Path):
    char_distinctive = _distinctive_terms(
        blueprint.get("visual_cast", []), "character_id", "locked_prompt_fragment"
    )
    loc_distinctive = _distinctive_terms(
        blueprint.get("locations", []), "location_id", "locked_prompt_fragment"
    )
    sheet_index = _build_sheet_index(visual_sheets_dir)

    data = {"scenes": []}
    for s in scenes:
        cuts = []
        for i, prompt in enumerate(s["cuts"], start=1):
            character_refs, location_ref, notes = _classify_cut_references(
                prompt, char_distinctive, loc_distinctive
            )
            reference_sheet_files = [
                sheet_index[cid] for cid in character_refs if cid in sheet_index
            ]
            if location_ref and location_ref in sheet_index:
                reference_sheet_files.append(sheet_index[location_ref])
            cuts.append(
                {
                    "cut_id": f"{s['scene_id']}-{i:02d}",
                    "image_prompt": prompt,
                    "character_refs": character_refs,
                    "location_ref": location_ref,
                    "reference_sheet_files": reference_sheet_files,
                    "notes": notes,
                }
            )
        data["scenes"].append({"scene_id": s["scene_id"], "zone_id": s["zone_id"], "cuts": cuts})

    (out_dir / "image_prompts_final.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return data


SCENE_BLOCK_FILLS = ("FFFFFF", "F2F2F2")


def write_image_generation_worklist(image_prompts_data: dict, xlsx_path: Path):
    """image_prompts_final.json相当のデータから、画像生成の手作業用Excelを作る"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "worklist"

    headers = [
        "scene_id",
        "cut_id",
        "zone_id",
        "画像No",
        "reference_sheet_files",
        "image_prompt",
        "notes",
        "生成済み",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    scene_top_border = Border(top=Side(style="medium", color="444444"))
    no_wrap_top = Alignment(wrap_text=False, vertical="top")

    row = 2
    seq = 1
    prev_scene_id = None
    fill_toggle = 0
    for scene in image_prompts_data["scenes"]:
        if scene["scene_id"] != prev_scene_id:
            fill_toggle = 1 - fill_toggle
        fill = PatternFill(
            fill_type="solid", start_color=SCENE_BLOCK_FILLS[fill_toggle], end_color=SCENE_BLOCK_FILLS[fill_toggle]
        )
        for cut in scene["cuts"]:
            is_scene_start = scene["scene_id"] != prev_scene_id
            values = [
                scene["scene_id"],
                cut["cut_id"],
                scene["zone_id"],
                seq,
                ", ".join(cut["reference_sheet_files"]),
                cut["image_prompt"],
                cut["notes"],
                "",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = fill
                cell.alignment = no_wrap_top
                if is_scene_start:
                    cell.border = scene_top_border
            prev_scene_id = scene["scene_id"]
            seq += 1
            row += 1

    widths = {1: 10, 2: 10, 3: 22, 4: 8, 5: 32, 6: 100, 7: 32, 8: 10}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(xlsx_path)


def _keywords(text: str):
    return [p for p in re.split(r"[・、,／/\s]+", text) if p]


def build_metadata_draft(blueprint: dict) -> dict:
    theme = blueprint.get("theme", "")
    focus = blueprint.get("explanation_focus", "")

    title_candidates = [
        f"{theme}で人生が狂った話",
        f"【実録シミュレーション】{theme}の末路",
        f"{focus}を知らずに払った代償——{theme}",
    ]

    plot_summary = blueprint.get("plot_summary", "")
    description = (
        f"{plot_summary}\n\n"
        f"この動画では「{theme}」をテーマに、判断ミスがどのように積み重なっていくかを再現しています。\n"
        f"予防策として「{focus}」についても解説しています。"
    ).strip()

    tags = []
    for text in (theme, focus):
        if text:
            tags.append(text)
            tags.extend(_keywords(text))
    tags = list(dict.fromkeys(tags))

    return {"title_candidates": title_candidates, "description": description, "tags": tags}


def write_metadata_draft(blueprint: dict, out_dir: Path):
    data = build_metadata_draft(blueprint)
    (out_dir / "metadata_draft.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


QUOTE_RE = re.compile(r"「(.*?)」")

# DRAFT値・暫定リスト。人物名の表記ゆれを網羅したものではない。
# 話者推測は前後1文の文脈一致による簡易ヒューリスティックであり、
# 誤判定を前提として「不明（要確認）」への逃がしを優先する設計とする
SPEAKER_KEYWORDS = {
    "健一": ["藤田健一", "健一"],
    "美和子": ["美和子", "妻"],
    "銀行員": ["行員"],
}

# visual_castに登録が無い脇役ラベルについて、image_prompt本文中の役割語から
# 該当CUTを検索するためのキーワード（DRAFT値・暫定リスト）
ROLE_SEARCH_TERMS = {
    "銀行員": ["bank teller"],
}

_GENDER_PATTERNS = [
    ("女性", re.compile(r"\b(woman|female)\b", re.IGNORECASE)),
    ("男性", re.compile(r"\b(man|male)\b", re.IGNORECASE)),
]


def _detect_gender(text: str):
    for label, pattern in _GENDER_PATTERNS:
        if pattern.search(text):
            return label
    return None


def _find_visual_cast_entry(label: str, blueprint: dict):
    for entry in blueprint.get("visual_cast", []):
        name = entry.get("name", "")
        if label in name or name in label:
            return entry
    return None


def _find_role_gender(scenes, terms):
    for scene in scenes:
        for cut_text in scene["cuts"]:
            if any(term.lower() in cut_text.lower() for term in terms):
                gender = _detect_gender(cut_text)
                if gender:
                    return scene["scene_id"], gender
    return None, None


def build_speaker_gender_summary(scenes, blueprint: dict) -> list:
    lines = ["## 話者一覧（画像プロンプトから抽出した性別情報）"]
    for label in SPEAKER_KEYWORDS:
        cast_entry = _find_visual_cast_entry(label, blueprint)
        if cast_entry:
            gender = _detect_gender(cast_entry.get("locked_prompt_fragment", "")) or "不明（要確認）"
            lines.append(f"- {label}：{gender}（{cast_entry['character_id']}、visual_cast参照）")
            continue
        terms = ROLE_SEARCH_TERMS.get(label)
        scene_id, gender = _find_role_gender(scenes, terms) if terms else (None, None)
        if gender:
            lines.append(f"- {label}：{gender}（SCENE {scene_id} CUT該当箇所の記述より）")
        else:
            lines.append(f"- {label}：不明（要確認）")
    lines.append("")
    return lines


def _split_narration_dialogue(narration: str):
    parts = []
    pos = 0
    for m in QUOTE_RE.finditer(narration):
        before = narration[pos:m.start()]
        if before:
            parts.append(("narration", before))
        parts.append(("dialogue", m.group(1)))
        pos = m.end()
    tail = narration[pos:]
    if tail:
        parts.append(("narration", tail))
    return parts


def _nearest_narration(parts, index, step):
    j = index + step
    while 0 <= j < len(parts):
        kind, text = parts[j]
        if kind == "narration":
            return text
        j += step
    return ""


def _last_sentence(text: str) -> str:
    for s in reversed(text.split("。")):
        if s.strip():
            return s
    return ""


def _first_sentence(text: str) -> str:
    for s in text.split("。"):
        if s.strip():
            return s
    return ""


def _match_speaker_labels(text: str):
    labels = set()
    for label, keywords in SPEAKER_KEYWORDS.items():
        if any(kw in text for kw in keywords):
            labels.add(label)
    return labels


def _guess_speaker(before_text: str, after_text: str):
    before_labels = _match_speaker_labels(_last_sentence(before_text))
    if len(before_labels) == 1:
        return next(iter(before_labels))
    after_labels = _match_speaker_labels(_first_sentence(after_text))
    if len(after_labels) == 1:
        return next(iter(after_labels))
    return None


def extract_dialogue_breakdown(scenes, blueprint: dict) -> str:
    lines = [
        "これは機械推測によるドラフトです。人間による目視確認・修正が必要です。",
        "",
    ]
    lines.extend(build_speaker_gender_summary(scenes, blueprint))
    for scene in scenes:
        lines.append(f"## SCENE {scene['scene_id']}")
        parts = _split_narration_dialogue(scene["narration"])
        for i, (kind, text) in enumerate(parts):
            if kind == "narration":
                lines.append(f"[ナレーション]: {text}")
            else:
                before_text = _nearest_narration(parts, i, -1)
                after_text = _nearest_narration(parts, i, 1)
                speaker = _guess_speaker(before_text, after_text)
                label = f"{speaker}セリフ" if speaker else "不明（要確認）"
                lines.append(f"[{label}]: {text}")
        lines.append("")
    return "\n".join(lines).rstrip("\n") + "\n"


def write_dialogue_breakdown(scenes, blueprint: dict, video_dir: Path):
    text = extract_dialogue_breakdown(scenes, blueprint)
    (video_dir / "voice_speaker_breakdown.md").write_text(text, encoding="utf-8")


def write_thumbnail_prompts(blueprint: dict, out_dir: Path):
    directions = blueprint.get("thumbnail_direction", [])[:3]
    data = {
        "thumbnail_prompts": [
            {"direction_ja": d, "prompt_draft": d, "needs_refinement": True} for d in directions
        ]
    }
    (out_dir / "thumbnail_prompts.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def main():
    parser = argparse.ArgumentParser(
        description="FINAL_SCRIPT_FULL.mdから制作データ（ナレーション・画像プロンプト・メタデータ叩き台）を抽出する"
    )
    parser.add_argument("script_path", type=Path, help="videos/{video_id}/FINAL_SCRIPT_FULL.md")
    args = parser.parse_args()

    script_path = args.script_path.resolve()
    if not script_path.exists():
        sys.exit(f"[ERROR] ファイルが見つかりません: {script_path}")

    video_dir = script_path.parent
    repo_root = Path(__file__).resolve().parent.parent
    blueprint_path = video_dir / "blueprint.json"
    genre_path = repo_root / "genres" / "risk_sim.json"
    metrics_path = repo_root / "metrics.csv"
    out_dir = video_dir / "output"

    if not blueprint_path.exists():
        sys.exit(f"[ERROR] blueprint.jsonが見つかりません: {blueprint_path}")
    blueprint = json.loads(blueprint_path.read_text(encoding="utf-8"))

    meta, scenes = parse_script(script_path.read_text(encoding="utf-8"))

    warnings = []
    validate_total_scenes(meta, scenes, warnings.append)
    validate_zone_char_budgets(meta, scenes, genre_path, warnings.append)
    validate_cut_pacing(meta, scenes, genre_path, warnings.append)
    validate_image_prompts(scenes, warnings.append)
    validate_policy_risk(scenes, warnings.append)
    validate_metrics_columns(metrics_path, warnings.append)

    out_dir.mkdir(parents=True, exist_ok=True)
    write_voice_plain_text(scenes, out_dir)
    write_voice_assignment_csv(scenes, out_dir)
    image_prompts_data = write_image_prompts(scenes, blueprint, video_dir / "visual_sheets", out_dir)
    write_image_generation_worklist(image_prompts_data, video_dir / "image_generation_worklist.xlsx")
    write_metadata_draft(blueprint, out_dir)
    write_thumbnail_prompts(blueprint, out_dir)
    write_dialogue_breakdown(scenes, blueprint, video_dir)

    for w in warnings:
        print(f"[WARNING] {w}", file=sys.stderr)
    print(f"抽出完了: {out_dir}")


if __name__ == "__main__":
    main()
